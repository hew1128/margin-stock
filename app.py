from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify
import sqlite3
import os
import re
import json
import tempfile
import io
from datetime import datetime
import xlrd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'margin_inventory_2024'

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get('DB_PATH', os.path.join(BASE_DIR, 'margin.db'))


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True) if os.path.dirname(DB_PATH) else None
    conn = get_db()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS products (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        option_name TEXT,
        sale_price INTEGER NOT NULL,
        purchase_price_cny REAL NOT NULL,
        exchange_rate REAL DEFAULT 190,
        customs_total INTEGER DEFAULT 0,
        shipping_total INTEGER DEFAULT 0,
        yongdal_total INTEGER DEFAULT 0,
        import_quantity INTEGER DEFAULT 1,
        naver_fee_rate REAL DEFAULT 2.0,
        domestic_shipping INTEGER DEFAULT 2500,
        is_active INTEGER DEFAULT 1,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    try:
        c.execute("ALTER TABLE products ADD COLUMN exchange_rate REAL DEFAULT 190")
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE products ADD COLUMN product_group TEXT")
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE products ADD COLUMN stock_group_id INTEGER")
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE products ADD COLUMN purchase_date TEXT")
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE products ADD COLUMN payment_method TEXT DEFAULT '위안화'")
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE products ADD COLUMN payment_card_info TEXT")
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE products ADD COLUMN purchase_price_krw INTEGER DEFAULT 0")
    except Exception:
        pass
    c.execute('''CREATE TABLE IF NOT EXISTS product_keywords (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        product_id INTEGER NOT NULL,
        keyword TEXT NOT NULL,
        FOREIGN KEY (product_id) REFERENCES products(id)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS stock_in (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        product_id INTEGER NOT NULL,
        quantity INTEGER NOT NULL,
        exchange_rate REAL NOT NULL,
        date TEXT NOT NULL,
        memo TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (product_id) REFERENCES products(id)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS stock_out (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        product_id INTEGER NOT NULL,
        quantity INTEGER NOT NULL,
        date TEXT NOT NULL,
        source TEXT DEFAULT 'manual',
        file_name TEXT,
        order_number TEXT,
        memo TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (product_id) REFERENCES products(id)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS product_option_defs (
        product_group TEXT PRIMARY KEY,
        basic_options TEXT NOT NULL DEFAULT '[]',
        addon_options TEXT NOT NULL DEFAULT '[]'
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS product_freebies (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        product_group TEXT NOT NULL,
        name TEXT NOT NULL,
        qty_per_order INTEGER DEFAULT 1,
        unit_cost INTEGER DEFAULT 0,
        stock INTEGER DEFAULT 0
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS freebie_rules (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        product_group TEXT NOT NULL,
        rule_name TEXT NOT NULL,
        freebie_id INTEGER NOT NULL,
        freebie_qty INTEGER DEFAULT 1,
        cond_type TEXT NOT NULL DEFAULT 'qty',
        min_qty INTEGER DEFAULT 1,
        option_name TEXT,
        option_value TEXT,
        is_active INTEGER DEFAULT 1,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    try:
        c.execute("ALTER TABLE products ADD COLUMN option_combo_json TEXT")
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE products ADD COLUMN total_import_qty INTEGER DEFAULT 0")
    except Exception:
        pass
    c.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('exchange_rate', '190')")
    # stock_in 없는 기존 상품 → import_quantity로 초기 입고 자동 생성
    # 단, 재고 공유 그룹의 비-마스터 상품(stock_group_id != id)은 제외
    no_stock = c.execute("""
        SELECT id, import_quantity, exchange_rate, created_at
        FROM products
        WHERE is_active=1
          AND id NOT IN (SELECT DISTINCT product_id FROM stock_in)
          AND (stock_group_id IS NULL OR stock_group_id = id)
    """).fetchall()
    for p in no_stock:
        date_str = (p['created_at'] or '')[:10] or datetime.now().strftime('%Y-%m-%d')
        c.execute("INSERT INTO stock_in (product_id, quantity, exchange_rate, date, memo) VALUES (?,?,?,?,?)",
            (p['id'], p['import_quantity'], p['exchange_rate'], date_str, '자동 초기 입고'))
    conn.commit()
    conn.close()


init_db()


def get_setting(key, default=None):
    conn = get_db()
    row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    conn.close()
    return row['value'] if row else default


def sort_groups_by_setting(group_keys, max_id_map=None):
    """저장된 그룹 순서 우선 적용, 없는 그룹은 뒤에 기본 순서"""
    stored = get_setting('group_order')
    order_map = {}
    if stored:
        try:
            for i, name in enumerate(json.loads(stored)):
                order_map[name] = i
        except Exception:
            pass
    def key_fn(g):
        if g in order_map:
            return (0, order_map[g])
        return (1, -(max_id_map.get(g, 0) if max_id_map else 0))
    return sorted(group_keys, key=key_fn)


@app.route('/group-order', methods=['POST'])
def update_group_order():
    data = request.get_json()
    order = data.get('order', [])
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('group_order', ?)",
                 (json.dumps(order, ensure_ascii=False),))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


def calculate_margin(product, exchange_rate=None, freebie_cost=0):
    keys = product.keys() if hasattr(product, 'keys') else product
    method = (product['payment_method'] if 'payment_method' in keys else None) or '위안화'
    krw_price = (product['purchase_price_krw'] if 'purchase_price_krw' in keys else 0) or 0
    sale = product['sale_price']
    if method in ('원화', '카드') and krw_price > 0:
        purchase_krw = krw_price
    else:
        rate = float(exchange_rate or product['exchange_rate'] or 190)
        purchase_krw = product['purchase_price_cny'] * rate
    customs_total = product['customs_total'] + product['shipping_total'] + product['yongdal_total']
    # 다중옵션 배치는 total_import_qty(전체 수량)로 개당 통관비 계산
    total_qty = (product['total_import_qty'] if 'total_import_qty' in keys and product['total_import_qty'] else None) \
                or product['import_quantity'] or 1
    customs_per_unit = customs_total / total_qty
    naver_fee = sale * product['naver_fee_rate'] / 100
    tax_amount = sale * 0.1
    total_cost_simple  = purchase_krw + customs_per_unit + naver_fee + product['domestic_shipping'] + freebie_cost
    total_cost_general = total_cost_simple + tax_amount
    simple_margin  = sale - total_cost_simple
    general_margin = sale - total_cost_general
    return {
        'purchase_krw':         round(purchase_krw, 1),
        'customs_total':        customs_total,
        'customs_per_unit':     round(customs_per_unit, 1),
        'naver_fee':            round(naver_fee, 1),
        'tax_amount':           round(tax_amount, 1),
        'freebie_cost':         round(freebie_cost, 1),
        'total_cost_simple':    round(total_cost_simple, 1),
        'total_cost_general':   round(total_cost_general, 1),
        'simple_margin':        round(simple_margin, 1),
        'simple_margin_rate':   round(simple_margin / sale * 100, 1) if sale else 0,
        'general_margin':       round(general_margin, 1),
        'general_margin_rate':  round(general_margin / sale * 100, 1) if sale else 0,
    }


def load_freebie_data():
    """사은품 정보 로드: {product_group: [freebie_dict, ...]}, {product_group: total_cost}"""
    conn = get_db()
    rows = conn.execute("SELECT * FROM product_freebies ORDER BY id").fetchall()
    conn.close()
    by_group = {}
    costs = {}
    for f in rows:
        pg = f['product_group']
        if pg not in by_group:
            by_group[pg] = []
            costs[pg] = 0
        by_group[pg].append(dict(f))
        costs[pg] += f['qty_per_order'] * f['unit_cost']
    return by_group, costs


def load_option_defs():
    """옵션 정의 로드: {product_group: {basic: [...], addon: [...]}}"""
    conn = get_db()
    rows = conn.execute("SELECT * FROM product_option_defs").fetchall()
    conn.close()
    result = {}
    for row in rows:
        result[row['product_group']] = {
            'basic': json.loads(row['basic_options'] or '[]'),
            'addon': json.loads(row['addon_options'] or '[]'),
        }
    return result


def apply_freebie_rules(conn, product_id, quantity, option_str):
    """주문 건에 사은품 규칙 적용 → 사은품 재고 차감. conn 미닫힘/commit 미호출."""
    product = conn.execute(
        "SELECT product_group, name, option_name FROM products WHERE id=?", (product_id,)).fetchone()
    if not product:
        return []
    pg = (product['product_group'] or '').strip() or product['name']
    rules = conn.execute("""
        SELECT fr.*, pf.name as freebie_name
        FROM freebie_rules fr
        JOIN product_freebies pf ON fr.freebie_id = pf.id
        WHERE fr.product_group=? AND fr.is_active=1
    """, (pg,)).fetchall()
    deducted = []
    for rule in rules:
        ctype   = rule['cond_type']
        qty_ok  = quantity >= rule['min_qty']
        opt_ok  = True
        if ctype in ('option', 'qty_and_option') and rule['option_value']:
            ov = rule['option_value'].strip()
            opt_ok = (ov in (option_str or '')) or (ov in (product['option_name'] or ''))
        triggered = (
            (ctype == 'qty'             and qty_ok) or
            (ctype == 'option'          and opt_ok) or
            (ctype == 'qty_and_option'  and qty_ok and opt_ok)
        )
        if triggered:
            conn.execute("UPDATE product_freebies SET stock=MAX(0, stock-?) WHERE id=?",
                         (rule['freebie_qty'], rule['freebie_id']))
            deducted.append({'rule_name': rule['rule_name'],
                             'freebie_name': rule['freebie_name'],
                             'qty': rule['freebie_qty']})
    return deducted


def get_current_stock(product_id):
    conn = get_db()
    row = conn.execute("SELECT stock_group_id FROM products WHERE id=?", (product_id,)).fetchone()
    gid = row['stock_group_id'] if row else None
    if gid:
        peers = [r['id'] for r in conn.execute(
            "SELECT id FROM products WHERE stock_group_id=?", (gid,)).fetchall()]
        if not peers:
            peers = [product_id]
        ph = ','.join('?' * len(peers))
        in_qty  = conn.execute(f"SELECT COALESCE(SUM(quantity),0) as t FROM stock_in  WHERE product_id IN ({ph})", peers).fetchone()['t']
        out_qty = conn.execute(f"SELECT COALESCE(SUM(quantity),0) as t FROM stock_out WHERE product_id IN ({ph})", peers).fetchone()['t']
    else:
        in_qty  = conn.execute("SELECT COALESCE(SUM(quantity),0) as t FROM stock_in  WHERE product_id=?", (product_id,)).fetchone()['t']
        out_qty = conn.execute("SELECT COALESCE(SUM(quantity),0) as t FROM stock_out WHERE product_id=?", (product_id,)).fetchone()['t']
    conn.close()
    return in_qty - out_qty


def get_all_products_with_keywords():
    conn = get_db()
    rows = conn.execute("""
        SELECT p.*, GROUP_CONCAT(pk.keyword, '||') as kws
        FROM products p
        LEFT JOIN product_keywords pk ON p.id = pk.product_id
        WHERE p.is_active = 1
        GROUP BY p.id ORDER BY p.id DESC
    """).fetchall()
    conn.close()
    return rows


# ─── 대시보드 ───────────────────────────────────────────────────────────────────

@app.route('/')
def dashboard():
    products = get_all_products_with_keywords()
    _, freebie_costs = load_freebie_data()
    items = []
    for p in products:
        pg = (p['product_group'] or '').strip() or p['name']
        fc = freebie_costs.get(pg, 0)
        combo = None
        if p['option_combo_json'] if 'option_combo_json' in p.keys() else None:
            try: combo = json.loads(p['option_combo_json'])
            except: pass
        items.append({'p': p, 'm': calculate_margin(p, freebie_cost=fc), 's': get_current_stock(p['id']),
                      'sgid': p['stock_group_id'] or p['id'], 'combo': combo})

    # product_group(또는 name) → (name, opt) 키 3단계 그룹핑
    group_map  = {}
    group_max_id = {}
    for item in items:
        pg   = (item['p']['product_group'] or '').strip()
        name = item['p']['name']
        opt  = item['p']['option_name'] or ''
        pid  = item['p']['id']
        gkey = pg if pg else name
        tkey = (name, opt)
        if gkey not in group_map:
            group_map[gkey] = {}
            group_max_id[gkey] = 0
        if tkey not in group_map[gkey]:
            group_map[gkey][tkey] = []
        group_map[gkey][tkey].append(item)
        if pid > group_max_id[gkey]:
            group_max_id[gkey] = pid

    opt_defs = load_option_defs()
    freebies_by_group, _ = load_freebie_data()
    gkeys = sort_groups_by_setting(list(group_map.keys()), group_max_id)
    grouped = []
    for gkey in gkeys:
        od = group_map[gkey]
        key_order = sorted(od.keys(), key=lambda k: max(i['p']['id'] for i in od[k]), reverse=True)
        opt_rows = []
        seen_sgids = set()
        total_g = 0
        for tkey in key_order:
            its = od[tkey]
            opt_s = sum(i['s'] for i in its)
            for i in its:
                if i['sgid'] not in seen_sgids:
                    seen_sgids.add(i['sgid'])
                    total_g += i['s']
            opt_rows.append({'opt': tkey[1], 'listing': tkey[0], 'latest': its[0], 'total_s': opt_s})
        grouped.append((gkey, opt_rows, total_g))
    return render_template('dashboard.html', grouped=grouped, group_names=gkeys,
                           opt_defs=opt_defs, freebies_by_group=freebies_by_group)


# ─── 상품 관리 ──────────────────────────────────────────────────────────────────

@app.route('/products')
def products():
    prods = get_all_products_with_keywords()
    _, freebie_costs = load_freebie_data()
    opt_defs = load_option_defs()
    freebies_by_group, _ = load_freebie_data()
    conn_r = get_db()
    rules_count_by_group = {r['product_group']: r['cnt'] for r in conn_r.execute(
        "SELECT product_group, COUNT(*) as cnt FROM freebie_rules WHERE is_active=1 GROUP BY product_group"
    ).fetchall()}
    conn_r.close()
    data = []
    for p in prods:
        pg = (p['product_group'] or '').strip() or p['name']
        fc = freebie_costs.get(pg, 0)
        combo = None
        if p['option_combo_json'] if 'option_combo_json' in p.keys() else None:
            try: combo = json.loads(p['option_combo_json'])
            except: pass
        max_addon_m = 0
        for ag in opt_defs.get(pg, {}).get('addon', []):
            best = max((ai.get('sell_price', 0) - ai.get('buy_price', 0) for ai in ag.get('items', [])), default=0)
            if best > 0:
                max_addon_m += best
        data.append({'p': p, 'm': calculate_margin(p, freebie_cost=fc), 's': get_current_stock(p['id']),
                     'sgid': p['stock_group_id'] or p['id'], 'combo': combo, 'max_addon_m': max_addon_m})
    # product_group(또는 name) → (name, opt) 키 3단계 그룹핑
    group_map  = {}
    group_max_id = {}
    for item in data:
        pg   = (item['p']['product_group'] or '').strip()
        name = item['p']['name']
        opt  = item['p']['option_name'] or ''
        pid  = item['p']['id']
        gkey = pg if pg else name
        tkey = (name, opt)
        if gkey not in group_map:
            group_map[gkey] = {}
            group_max_id[gkey] = 0
        if tkey not in group_map[gkey]:
            group_map[gkey][tkey] = []
        group_map[gkey][tkey].append(item)
        if pid > group_max_id[gkey]:
            group_max_id[gkey] = pid

    gkeys = sort_groups_by_setting(list(group_map.keys()), group_max_id)
    grouped = []
    for gkey in gkeys:
        od = group_map[gkey]
        key_order = sorted(od.keys(), key=lambda k: max(i['p']['id'] for i in od[k]), reverse=True)
        opt_list = [{'opt': k[1], 'listing': k[0], 'batches': od[k]} for k in key_order]
        seen_sgids = set()
        group_stock = 0
        for row in opt_list:
            for item in row['batches']:
                if item['sgid'] not in seen_sgids:
                    seen_sgids.add(item['sgid'])
                    group_stock += item['s']
        grouped.append((gkey, opt_list, group_stock))
    return render_template('products.html', grouped=grouped, group_names=gkeys,
                           opt_defs=opt_defs, freebies_by_group=freebies_by_group,
                           rules_count_by_group=rules_count_by_group)


@app.route('/products/new', methods=['GET', 'POST'])
def product_new():
    if request.method == 'POST':
        conn = get_db()
        c = conn.cursor()
        # 공통 설정
        pg        = request.form.get('product_group', '')
        customs   = int(request.form.get('customs_total') or 0)
        shipping  = int(request.form.get('shipping_total') or 0)
        yongdal   = int(request.form.get('yongdal_total') or 0)
        nfr       = float(request.form.get('naver_fee_rate') or 2.0)
        dom_ship  = int(request.form.get('domestic_shipping') or 2500)
        today     = datetime.now().strftime('%Y-%m-%d')
        # 묶음별 데이터 (입고 개수, 사입가, 환율, 사입날짜, 결제방식은 묶음 레벨)
        group_qtys    = request.form.getlist('group_qtys[]')
        group_cnys    = request.form.getlist('group_cnys[]')
        group_rates   = request.form.getlist('group_rates[]')
        group_dates   = request.form.getlist('group_dates[]')
        group_methods = request.form.getlist('group_methods[]')
        group_cards   = request.form.getlist('group_cards[]')
        group_krws    = request.form.getlist('group_krws[]')
        # 행 단위 데이터
        names   = request.form.getlist('names[]')
        opts    = request.form.getlist('option_names[]')
        prices  = request.form.getlist('sale_prices[]')
        gidxs   = request.form.getlist('group_indices[]')

        # 묶음 인덱스 → [product_id] 매핑
        groups = {}
        for i, name in enumerate(names):
            if not name.strip():
                continue
            gidx  = int(gidxs[i]) if i < len(gidxs) and gidxs[i] != '' else 0
            opt   = opts[i].strip() if i < len(opts) else ''
            price = int(prices[i])  if i < len(prices) and prices[i] else 0
            qty   = int(group_qtys[gidx])   if gidx < len(group_qtys)  and group_qtys[gidx]  else 1
            pdate = group_dates[gidx]        if gidx < len(group_dates) and group_dates[gidx] else today
            method    = group_methods[gidx]  if gidx < len(group_methods) and group_methods[gidx] else '위안화'
            card_info = group_cards[gidx]    if gidx < len(group_cards)  and group_cards[gidx]  else ''
            krw_price = int(group_krws[gidx]) if gidx < len(group_krws) and group_krws[gidx] else 0
            if method in ('원화', '카드'):
                cny  = 0.0
                rate = 1.0
            else:
                cny  = float(group_cnys[gidx])  if gidx < len(group_cnys)  and group_cnys[gidx]  else 0
                rate = float(group_rates[gidx]) if gidx < len(group_rates) and group_rates[gidx] else 190
            c.execute("""INSERT INTO products
                (name, option_name, product_group, sale_price, purchase_price_cny, exchange_rate,
                 customs_total, shipping_total, yongdal_total, import_quantity,
                 naver_fee_rate, domestic_shipping, purchase_date,
                 payment_method, payment_card_info, purchase_price_krw)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (name.strip(), opt, pg, price, cny, rate,
                 customs, shipping, yongdal, qty, nfr, dom_ship, pdate,
                 method, card_info, krw_price))
            pid = c.lastrowid
            if gidx not in groups:
                groups[gidx] = []
            groups[gidx].append(pid)

        # 묶음별 stock_group_id 설정 + stock_in (마스터 1개만)
        count = 0
        for gidx, pids in groups.items():
            qty   = int(group_qtys[gidx])   if gidx < len(group_qtys)  and group_qtys[gidx]  else 1
            rate  = float(group_rates[gidx]) if gidx < len(group_rates) and group_rates[gidx] else 190
            pdate = group_dates[gidx] if gidx < len(group_dates) and group_dates[gidx] else today
            master = pids[0]
            for pid in pids:
                c.execute("UPDATE products SET stock_group_id=? WHERE id=?", (master, pid))
            c.execute("INSERT INTO stock_in (product_id, quantity, exchange_rate, date, memo) VALUES (?,?,?,?,?)",
                (master, qty, rate, pdate, '상품 등록 초기 입고'))
            count += len(pids)
        conn.commit()
        conn.close()
        flash(f'{count}개 상품이 등록되었습니다.')
        return redirect(url_for('products'))
    # 배치 추가 모드: ?group=NAME 으로 진입 시 기존 묶음 구조 pre-fill
    group_name   = request.args.get('group', '')
    batch_groups = []
    common       = None
    is_add_batch = False

    if group_name:
        conn = get_db()
        all_ps = conn.execute(
            "SELECT * FROM products WHERE product_group=? AND is_active=1 ORDER BY id",
            (group_name,)
        ).fetchall()
        conn.close()
        if all_ps:
            sg_map = {}
            for p in all_ps:
                sgid = p['stock_group_id'] or p['id']
                if sgid not in sg_map:
                    sg_map[sgid] = []
                sg_map[sgid].append(p)
            # 고유 option tier별 최신 배치만
            fp_latest = {}
            for sgid in sorted(sg_map.keys()):
                prods = sg_map[sgid]
                fp = frozenset((p['name'], p['option_name'] or '') for p in prods)
                fp_latest[fp] = (sgid, prods)
            for _fp, (sgid, prods) in sorted(fp_latest.items(), key=lambda x: x[1][0]):
                listings = [{'name': p['name'], 'option': p['option_name'] or '', 'sale_price': p['sale_price']}
                            for p in sorted(prods, key=lambda x: x['id'])]
                batch_groups.append(listings)
            latest_p = max(all_ps, key=lambda x: x['id'])
            common = {k: latest_p[k] for k in ('customs_total','shipping_total','yongdal_total','naver_fee_rate','domestic_shipping')}
            is_add_batch = True

    return render_template('product_form.html', product=None, keywords='',
                           is_add_batch=is_add_batch, group_name=group_name,
                           batch_groups=batch_groups, common=common)


@app.route('/products/<int:pid>/edit', methods=['GET', 'POST'])
def product_edit(pid):
    if request.method == 'POST':
        new_qty = int(request.form.get('import_quantity') or 1)
        new_rate = float(request.form.get('exchange_rate') or 190)
        conn = get_db()
        conn.execute("""UPDATE products SET
            name=?, option_name=?, product_group=?, sale_price=?, purchase_price_cny=?, exchange_rate=?,
            customs_total=?, shipping_total=?, yongdal_total=?, import_quantity=?,
            naver_fee_rate=?, domestic_shipping=?
            WHERE id=?""",
            (request.form['name'], request.form.get('option_name', ''),
             request.form.get('product_group', ''),
             int(request.form['sale_price']), float(request.form['purchase_price_cny']),
             new_rate,
             int(request.form.get('customs_total') or 0), int(request.form.get('shipping_total') or 0),
             int(request.form.get('yongdal_total') or 0), new_qty,
             float(request.form.get('naver_fee_rate') or 2.0), int(request.form.get('domestic_shipping') or 2500),
             pid))
        # 자동 생성된 초기 입고 수량도 import_quantity에 맞춰 업데이트
        conn.execute("""UPDATE stock_in SET quantity=?, exchange_rate=?
                        WHERE product_id=?
                          AND memo IN ('상품 등록 초기 입고', '자동 초기 입고')""",
                     (new_qty, new_rate, pid))
        conn.commit()
        conn.close()
        flash('상품이 수정되었습니다.')
        return redirect(url_for('products'))

    # GET
    conn = get_db()
    product = conn.execute("SELECT * FROM products WHERE id=?", (pid,)).fetchone()
    kws = conn.execute("SELECT keyword FROM product_keywords WHERE product_id=?", (pid,)).fetchall()
    keywords_text = '\n'.join(k['keyword'] for k in kws)
    conn.close()
    return render_template('product_form.html', product=product, keywords=keywords_text)


@app.route('/products/new-multi', methods=['GET', 'POST'])
def product_new_multi():
    if request.method == 'POST':
        conn = get_db()
        c = conn.cursor()
        pg = request.form.get('product_group', '').strip()
        if not pg:
            flash('제품명(그룹)을 입력해주세요.')
            conn.close()
            return redirect(request.url)

        # 기본 옵션 그룹 파싱
        basic_options = []
        for i in range(1, 4):
            name = request.form.get(f'basic_opt_name_{i}', '').strip()
            vals_str = request.form.get(f'basic_opt_values_{i}', '').strip()
            if name and vals_str:
                vals = [v.strip() for v in vals_str.split(',') if v.strip()]
                if vals:
                    basic_options.append({'name': name, 'values': vals})

        # 추가 옵션 파싱 (형식: 값:판매추가가:사입추가가)
        addon_options = []
        for i in range(1, 4):
            name = request.form.get(f'addon_opt_name_{i}', '').strip()
            items_str = request.form.get(f'addon_opt_items_{i}', '').strip()
            if name and items_str:
                items = []
                for part in items_str.split(','):
                    ps = [x.strip() for x in part.strip().split(':')]
                    if ps and ps[0]:
                        items.append({
                            'value':      ps[0],
                            'sell_price': int(ps[1]) if len(ps) > 1 and ps[1].isdigit() else 0,
                            'buy_price':  int(ps[2]) if len(ps) > 2 and ps[2].isdigit() else 0,
                        })
                if items:
                    addon_options.append({'name': name, 'items': items})

        # 옵션 정의 저장
        c.execute("INSERT OR REPLACE INTO product_option_defs (product_group, basic_options, addon_options) VALUES (?,?,?)",
                  (pg, json.dumps(basic_options, ensure_ascii=False),
                   json.dumps(addon_options, ensure_ascii=False)))

        # 사은품 저장
        for i in range(1, 4):
            fname = request.form.get(f'freebie_name_{i}', '').strip()
            if fname:
                qty_per  = int(request.form.get(f'freebie_qty_{i}')   or 1)
                unit_cost = int(request.form.get(f'freebie_cost_{i}')  or 0)
                stock    = int(request.form.get(f'freebie_stock_{i}')  or 0)
                c.execute("INSERT INTO product_freebies (product_group, name, qty_per_order, unit_cost, stock) VALUES (?,?,?,?,?)",
                          (pg, fname, qty_per, unit_cost, stock))

        # 공통 원가
        customs   = int(request.form.get('customs_total')    or 0)
        shipping  = int(request.form.get('shipping_total')   or 0)
        yongdal   = int(request.form.get('yongdal_total')    or 0)
        nfr       = float(request.form.get('naver_fee_rate') or 2.0)
        dom_ship  = int(request.form.get('domestic_shipping') or 2500)

        # 사입 정보
        pdate    = request.form.get('group_date') or datetime.now().strftime('%Y-%m-%d')
        method   = request.form.get('group_method', '위안화')
        cny      = float(request.form.get('group_cny')  or 0)
        rate     = float(request.form.get('group_rate') or 190)
        krw      = int(request.form.get('group_krw')    or 0)
        card_info = request.form.get('group_card', '').strip()
        if method in ('원화', '카드'):
            cny = 0.0; rate = 1.0
        else:
            krw = 0

        # SKU 데이터
        sku_combos = request.form.getlist('sku_combo[]')
        sku_prices = request.form.getlist('sku_sale_price[]')
        sku_qtys   = request.form.getlist('sku_qty[]')
        total_qty  = sum(int(q) for q in sku_qtys if q and q.isdigit())

        pids = []
        for i, combo_json in enumerate(sku_combos):
            try:
                combo = json.loads(combo_json)
            except Exception:
                continue
            sale_price = int(sku_prices[i]) if i < len(sku_prices) and sku_prices[i] else 0
            qty        = int(sku_qtys[i])   if i < len(sku_qtys)   and sku_qtys[i]   else 0
            option_name = ' / '.join(combo.values())
            c.execute("""INSERT INTO products
                (name, option_name, product_group, sale_price, purchase_price_cny, exchange_rate,
                 customs_total, shipping_total, yongdal_total, import_quantity,
                 naver_fee_rate, domestic_shipping, purchase_date,
                 payment_method, payment_card_info, purchase_price_krw,
                 option_combo_json, total_import_qty)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (pg, option_name, pg, sale_price, cny, rate,
                 customs, shipping, yongdal, qty, nfr, dom_ship, pdate,
                 method, card_info, krw, combo_json, total_qty))
            pid = c.lastrowid
            pids.append((pid, qty))

        # stock_group_id 설정 + stock_in
        for pid, qty in pids:
            c.execute("UPDATE products SET stock_group_id=? WHERE id=?", (pid, pid))
            if qty > 0:
                c.execute("INSERT INTO stock_in (product_id, quantity, exchange_rate, date, memo) VALUES (?,?,?,?,?)",
                          (pid, qty, rate, pdate, '다중옵션 등록 초기 입고'))

        conn.commit()
        conn.close()
        flash(f'{len(pids)}개 SKU 등록 완료.')
        return redirect(url_for('products'))

    return render_template('product_form_multi.html')


@app.route('/freebies/<int:fid>/adjust', methods=['POST'])
def freebie_adjust(fid):
    action = request.form.get('action', 'add')
    qty    = int(request.form.get('qty') or 0)
    conn = get_db()
    if action == 'add':
        conn.execute("UPDATE product_freebies SET stock=stock+? WHERE id=?", (qty, fid))
    elif action == 'subtract':
        conn.execute("UPDATE product_freebies SET stock=MAX(0, stock-?) WHERE id=?", (qty, fid))
    conn.commit()
    conn.close()
    return redirect(request.referrer or url_for('products'))


@app.route('/products/<group>/freebie-rules', methods=['GET', 'POST'])
def freebie_rules_page(group):
    if request.method == 'POST':
        rule_name    = request.form.get('rule_name', '').strip()
        cond_type    = request.form.get('cond_type', 'qty')
        min_qty      = int(request.form.get('min_qty') or 1)
        option_name  = request.form.get('option_name', '').strip()
        option_value = request.form.get('option_value', '').strip()
        freebie_id   = int(request.form.get('freebie_id') or 0)
        freebie_qty  = int(request.form.get('freebie_qty') or 1)
        if not rule_name or not freebie_id:
            flash('규칙 이름과 사은품을 입력해주세요.')
        else:
            conn = get_db()
            conn.execute("""INSERT INTO freebie_rules
                (product_group, rule_name, freebie_id, freebie_qty, cond_type, min_qty, option_name, option_value)
                VALUES (?,?,?,?,?,?,?,?)""",
                (group, rule_name, freebie_id, freebie_qty, cond_type, min_qty,
                 option_name  if cond_type != 'qty' else None,
                 option_value if cond_type != 'qty' else None))
            conn.commit()
            conn.close()
            flash('규칙이 추가되었습니다.')
        return redirect(url_for('freebie_rules_page', group=group))

    conn = get_db()
    freebies = [dict(r) for r in conn.execute(
        "SELECT * FROM product_freebies WHERE product_group=? ORDER BY id", (group,)).fetchall()]
    rules = [dict(r) for r in conn.execute("""
        SELECT fr.*, pf.name as freebie_name
        FROM freebie_rules fr
        LEFT JOIN product_freebies pf ON fr.freebie_id = pf.id
        WHERE fr.product_group=? AND fr.is_active=1
        ORDER BY fr.id
    """, (group,)).fetchall()]
    opt_def_row = conn.execute(
        "SELECT basic_options FROM product_option_defs WHERE product_group=?", (group,)).fetchone()
    conn.close()
    basic_options = json.loads(opt_def_row['basic_options'] if opt_def_row else '[]')
    return render_template('freebie_rules.html', group=group, freebies=freebies, rules=rules,
                           basic_options=basic_options,
                           basic_options_json=json.dumps(basic_options, ensure_ascii=False))


@app.route('/freebie-rules/<int:rid>/delete', methods=['POST'])
def freebie_rule_delete(rid):
    conn = get_db()
    rule = conn.execute("SELECT product_group FROM freebie_rules WHERE id=?", (rid,)).fetchone()
    group = rule['product_group'] if rule else None
    conn.execute("DELETE FROM freebie_rules WHERE id=?", (rid,))
    conn.commit()
    conn.close()
    flash('규칙이 삭제되었습니다.')
    return redirect(url_for('freebie_rules_page', group=group) if group else url_for('products'))


@app.route('/freebie-rules/<int:rid>/manual', methods=['POST'])
def freebie_rule_manual(rid):
    qty        = int(request.form.get('qty') or 1)
    option_str = request.form.get('option_str', '').strip()
    conn = get_db()
    rule = conn.execute("""
        SELECT fr.*, pf.name as freebie_name
        FROM freebie_rules fr
        LEFT JOIN product_freebies pf ON fr.freebie_id = pf.id
        WHERE fr.id=?
    """, (rid,)).fetchone()
    if rule:
        ctype  = rule['cond_type']
        qty_ok = qty >= rule['min_qty']
        opt_ok = True
        if ctype in ('option', 'qty_and_option') and rule['option_value']:
            ov = rule['option_value'].strip()
            opt_ok = (ov in option_str) if option_str else True
        triggered = (
            (ctype == 'qty'            and qty_ok) or
            (ctype == 'option'         and opt_ok) or
            (ctype == 'qty_and_option' and qty_ok and opt_ok)
        )
        group = rule['product_group']
        if triggered:
            conn.execute("UPDATE product_freebies SET stock=MAX(0, stock-?) WHERE id=?",
                         (rule['freebie_qty'], rule['freebie_id']))
            conn.commit()
            flash(f'[{rule["rule_name"]}] 사은품 "{rule["freebie_name"]}" {rule["freebie_qty"]}개 수동 차감 완료')
        else:
            flash(f'[{rule["rule_name"]}] 조건 미충족 (수량:{qty} / 옵션:"{option_str or "-"}") — 차감하지 않았습니다')
    conn.close()
    return redirect(request.referrer or url_for('products'))


@app.route('/products/<int:pid>/delete', methods=['POST'])
def product_delete(pid):
    conn = get_db()
    conn.execute("UPDATE products SET is_active=0 WHERE id=?", (pid,))
    conn.commit()
    conn.close()
    flash('상품이 삭제되었습니다.')
    return redirect(url_for('products'))


@app.route('/stock/add/<int:pid>', methods=['POST'])
def stock_add(pid):
    qty  = int(request.form.get('quantity') or 0)
    memo = request.form.get('memo', '').strip() or '수동 재고 추가'
    date = request.form.get('date') or datetime.now().strftime('%Y-%m-%d')
    if qty <= 0:
        flash('수량을 1개 이상 입력해주세요.')
        return redirect(url_for('products'))
    conn = get_db()
    rate = conn.execute("SELECT exchange_rate FROM products WHERE id=?", (pid,)).fetchone()['exchange_rate']
    conn.execute("INSERT INTO stock_in (product_id, quantity, exchange_rate, date, memo) VALUES (?,?,?,?,?)",
                 (pid, qty, rate, date, memo))
    conn.commit()
    conn.close()
    flash(f'{qty}개 재고가 추가되었습니다. ({memo})')
    return redirect(url_for('products'))


@app.route('/products/trash')
def products_trash():
    conn = get_db()
    rows = conn.execute("""
        SELECT p.*, GROUP_CONCAT(pk.keyword, '||') as kws
        FROM products p
        LEFT JOIN product_keywords pk ON p.id = pk.product_id
        WHERE p.is_active = 0
        GROUP BY p.id ORDER BY p.id DESC
    """).fetchall()
    conn.close()
    return render_template('products_trash.html', products=rows)


@app.route('/products/<int:pid>/restore', methods=['POST'])
def product_restore(pid):
    conn = get_db()
    conn.execute("UPDATE products SET is_active=1 WHERE id=?", (pid,))
    conn.commit()
    conn.close()
    flash('상품이 복구되었습니다.')
    return redirect(url_for('products_trash'))


# ─── 입고 등록 ──────────────────────────────────────────────────────────────────

@app.route('/stock/in', methods=['GET', 'POST'])
def stock_in():
    if request.method == 'POST':
        conn = get_db()
        conn.execute("INSERT INTO stock_in (product_id, quantity, exchange_rate, date, memo) VALUES (?,?,?,?,?)",
            (int(request.form['product_id']), int(request.form['quantity']),
             float(request.form['exchange_rate']), request.form['date'],
             request.form.get('memo', '')))
        conn.commit()
        conn.close()
        flash('입고가 등록되었습니다.')
        return redirect(url_for('dashboard'))

    products = get_all_products_with_keywords()
    exchange_rate = get_setting('exchange_rate', '190')
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('stock_in.html', products=products, exchange_rate=exchange_rate, today=today)


# ─── 발송파일 업로드 ────────────────────────────────────────────────────────────

@app.route('/stock/upload', methods=['GET', 'POST'])
def stock_upload():
    if request.method == 'POST':
        # 출고 확정
        if request.form.get('confirm') == '1':
            pending = session.pop('pending_out', [])
            filename = session.pop('pending_file', '')
            all_deducted = []
            if pending:
                today = datetime.now().strftime('%Y-%m-%d')
                conn = get_db()
                for r in pending:
                    conn.execute("INSERT INTO stock_out (product_id, quantity, date, source, file_name, order_number) VALUES (?,?,?,'file',?,?)",
                        (r['product_id'], r['quantity'], today, filename, r['order_number']))
                    all_deducted.extend(apply_freebie_rules(conn, r['product_id'], r['quantity'], r.get('option_str', '')))
                conn.commit()
                conn.close()
                msg = f'{len(pending)}건 출고 처리되었습니다.'
                if all_deducted:
                    qty_sum = {}
                    for d in all_deducted:
                        qty_sum[d['freebie_name']] = qty_sum.get(d['freebie_name'], 0) + d['qty']
                    msg += ' | 사은품 자동 차감: ' + ', '.join(f'{n} {q}개' for n, q in qty_sum.items())
                flash(msg)
            return redirect(url_for('dashboard'))

        # 파일 파싱
        file = request.files.get('file')
        if not file or not file.filename:
            flash('파일을 선택해주세요.')
            return redirect(request.url)

        suffix = '.xls' if file.filename.lower().endswith('.xls') else '.xlsx'
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        file.save(tmp.name)
        tmp.close()

        try:
            wb = xlrd.open_workbook(tmp.name)
            ws = wb.sheet_by_index(0)
            products = get_all_products_with_keywords()
            results, unmatched = [], []

            for ri in range(1, ws.nrows):
                g_val = str(ws.cell_value(ri, 6))   # G열: 상품명+옵션+수량
                j_val = str(ws.cell_value(ri, 9))   # J열: 배송메세지 (fallback)
                order = str(ws.cell_value(ri, 0))   # A열: 주문번호

                # G열에서 옵션명 추출: "옵션: 기본 연결형☞18EA"
                opt_m = re.search(r'옵션[:\s]*(.+?)☞', g_val)
                option_str = opt_m.group(1).strip() if opt_m else ''

                # G열에서 수량 추출: "☞18EA"
                qty_m = re.search(r'☞(\d+)\s*EA', g_val, re.IGNORECASE)
                if not qty_m:
                    # fallback: J열 "(총N개)"
                    qty_m = re.search(r'총(\d+)개', j_val)
                if not qty_m:
                    unmatched.append({'order': order, 'name': g_val[:50], 'reason': '수량 파싱 불가'})
                    continue

                qty = int(qty_m.group(1))

                # 상품명 기반 매칭 — p.name 이 G열에 포함되면 매칭, 옵션명도 일치하면 우선
                scored = []
                for p in products:
                    pname = (p['name'] or '').strip()
                    popt  = (p['option_name'] or '').strip()
                    if not pname or pname not in g_val:
                        continue
                    score = len(pname)  # 긴 이름일수록 구체적 매칭
                    if popt and option_str and popt in option_str:
                        score += 100000
                    elif popt and popt in g_val:
                        score += 50000
                    scored.append((score, p))
                candidates = []
                if scored:
                    max_score = max(s for s, _ in scored)
                    candidates = [p for s, p in scored if s == max_score]
                # FIFO: 최고점 후보 중 재고 있는 가장 오래된 배치 우선
                matched = None
                for c in sorted(candidates, key=lambda x: x['id']):
                    if get_current_stock(c['id']) > 0:
                        matched = c
                        break
                if not matched and candidates:
                    matched = min(candidates, key=lambda x: x['id'])

                if matched:
                    results.append({
                        'product_id': matched['id'],
                        'product_name': matched['name'],
                        'option_name': matched['option_name'] or '',
                        'quantity': qty,
                        'order_number': order,
                        'raw_name': g_val[:60],
                        'option_str': option_str,
                    })
                else:
                    unmatched.append({'order': order, 'name': f'{option_str or g_val[:40]}', 'reason': '상품 매칭 실패'})

            os.unlink(tmp.name)
            session['pending_out'] = results
            session['pending_file'] = file.filename
            return render_template('stock_upload.html', results=results, unmatched=unmatched, filename=file.filename)

        except Exception as e:
            os.unlink(tmp.name)
            flash(f'파일 파싱 오류: {str(e)}')
            return redirect(request.url)

    # GET
    if request.args.get('reset'):
        session.pop('pending_out', None)
        session.pop('pending_file', None)
        return render_template('stock_upload.html', results=None, unmatched=None, filename=None)
    pending = session.get('pending_out')
    if pending:
        return render_template('stock_upload.html', results=pending, unmatched=[], filename=session.get('pending_file', ''))
    return render_template('stock_upload.html', results=None, unmatched=None, filename=None)


@app.route('/stock/upload/cancel')
def stock_upload_cancel():
    session.pop('pending_out', None)
    session.pop('pending_file', None)
    return redirect(url_for('dashboard'))


# ─── 입출고 이력 ────────────────────────────────────────────────────────────────

@app.route('/history')
def history():
    conn = get_db()
    prods = get_all_products_with_keywords()

    result = []
    for p in prods:
        pid = p['id']
        total_in  = conn.execute("SELECT COALESCE(SUM(quantity),0) as t FROM stock_in  WHERE product_id=?", (pid,)).fetchone()['t']
        total_out = conn.execute("SELECT COALESCE(SUM(quantity),0) as t FROM stock_out WHERE product_id=?", (pid,)).fetchone()['t']
        ins  = conn.execute("SELECT * FROM stock_in  WHERE product_id=? ORDER BY date DESC, id DESC", (pid,)).fetchall()
        outs = conn.execute("SELECT * FROM stock_out WHERE product_id=? ORDER BY date DESC, id DESC", (pid,)).fetchall()
        if total_in > 0 or total_out > 0:
            result.append({'p': p, 'total_in': total_in, 'total_out': total_out,
                           'current': total_in - total_out, 'ins': ins, 'outs': outs})

    conn.close()

    # product_group 기준으로 그룹핑 (대시보드/상품관리와 동일)
    g_map = {}
    g_max_id = {}
    for item in result:
        pg   = (item['p']['product_group'] or '').strip()
        name = item['p']['name']
        gkey = pg if pg else name
        pid  = item['p']['id']
        if gkey not in g_map:
            g_map[gkey] = []
            g_max_id[gkey] = 0
        g_map[gkey].append(item)
        if pid > g_max_id[gkey]:
            g_max_id[gkey] = pid
    gkeys = sort_groups_by_setting(list(g_map.keys()), g_max_id)
    grouped = [(gkey, g_map[gkey]) for gkey in gkeys]
    return render_template('history.html', grouped=grouped, group_names=gkeys)


# ─── 분석 ───────────────────────────────────────────────────────────────────────

@app.route('/analytics')
def analytics():
    conn = get_db()
    rows = conn.execute("""
        SELECT COALESCE(NULLIF(p.product_group,''), p.name) as grp,
               p.name, COALESCE(p.option_name,'') as opt,
               so.date, SUM(so.quantity) as qty
        FROM stock_out so
        JOIN products p ON so.product_id = p.id
        GROUP BY grp, p.name, p.option_name, so.date
        ORDER BY grp, p.name, p.option_name, so.date
    """).fetchall()
    conn.close()

    data = {}
    group_max_date = {}
    for row in rows:
        grp  = row['grp']
        name = row['name']
        opt  = row['opt'] or '기본'
        if grp not in data:
            data[grp] = {}
            group_max_date[grp] = ''
        if name not in data[grp]:
            data[grp][name] = {}
        if opt not in data[grp][name]:
            data[grp][name][opt] = []
        data[grp][name][opt].append({'date': row['date'], 'qty': row['qty']})
        if row['date'] > group_max_date.get(grp, ''):
            group_max_date[grp] = row['date']

    # max_date를 숫자로 쓰기 위해 group_max_id가 없으므로 날짜역순을 fallback으로 사용
    def _date_to_pseudo_id(g):
        d = group_max_date.get(g, '')
        return int(d.replace('-', '')) if d else 0
    fallback_map = {g: _date_to_pseudo_id(g) for g in data}
    gkeys = sort_groups_by_setting(list(data.keys()), fallback_map)
    ordered = {g: data[g] for g in gkeys}
    return render_template('analytics.html',
                           chart_data=json.dumps(ordered, ensure_ascii=False),
                           group_names=gkeys)


# ─── 엑셀 다운로드 ──────────────────────────────────────────────────────────────

def _xl_header(ws, cols, fill_color='4F46E5'):
    """헤더 행 스타일 적용"""
    fill = PatternFill('solid', fgColor=fill_color)
    font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(bottom=Side(style='thin', color='CCCCCC'))
    for ci, (label, width) in enumerate(cols, 1):
        c = ws.cell(1, ci, label)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 22


def _xl_row(ws, row, values, alt=False):
    """데이터 행 스타일 적용"""
    bg = 'F9FAFB' if alt else 'FFFFFF'
    fill = PatternFill('solid', fgColor=bg)
    border = Border(bottom=Side(style='thin', color='F0F0F0'))
    for ci, val in enumerate(values, 1):
        c = ws.cell(row, ci, val)
        c.fill = fill
        c.border = border
        c.alignment = Alignment(vertical='center')
        if isinstance(val, (int, float)):
            c.alignment = Alignment(horizontal='right', vertical='center')


@app.route('/download/products')
def download_products():
    prods = get_all_products_with_keywords()
    data = [{'p': p, 'm': calculate_margin(p), 's': get_current_stock(p['id'])} for p in prods]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '상품관리'
    ws.freeze_panes = 'A2'

    cols = [
        ('상품명', 22), ('옵션명', 14),
        ('판매가(원)', 13), ('사입가(위안)', 13), ('환율', 8),
        ('관세합계', 12), ('운송비합계', 12), ('용달합계', 12), ('입고개수', 10),
        ('개당통관비', 12), ('수수료율(%)', 12), ('수수료(원)', 12), ('국내택배비', 12), ('일반세금(원)', 12),
        ('현재재고', 10),
        ('간이총원가', 13), ('간이마진(원)', 13), ('간이마진율(%)', 13),
        ('일반총원가', 13), ('일반마진(원)', 13), ('일반마진율(%)', 13),
        ('키워드', 30),
    ]
    _xl_header(ws, cols)

    for ri, item in enumerate(data, 2):
        p, m, s = item['p'], item['m'], item['s']
        kws = (p['kws'] or '').replace('||', ', ')
        _xl_row(ws, ri, [
            p['name'], p['option_name'] or '',
            p['sale_price'], p['purchase_price_cny'], p['exchange_rate'],
            p['customs_total'], p['shipping_total'], p['yongdal_total'], p['import_quantity'],
            round(m['customs_per_unit'], 1), p['naver_fee_rate'], round(m['naver_fee'], 1),
            p['domestic_shipping'], round(m['tax_amount'], 1),
            s,
            round(m['total_cost_simple'], 1), round(m['simple_margin'], 1), round(m['simple_margin_rate'], 1),
            round(m['total_cost_general'], 1), round(m['general_margin'], 1), round(m['general_margin_rate'], 1),
            kws,
        ], alt=(ri % 2 == 0))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"상품관리_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/download/stock')
def download_stock():
    prods = get_all_products_with_keywords()
    data = [{'p': p, 'm': calculate_margin(p), 's': get_current_stock(p['id'])} for p in prods]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '재고현황'
    ws.freeze_panes = 'A2'

    cols = [
        ('상품명', 22), ('옵션명', 14), ('판매가(원)', 13), ('사입가(위안)', 13),
        ('환율', 8), ('입고개수', 10), ('현재재고', 10),
        ('간이총원가', 13), ('간이마진(원)', 13), ('간이마진율(%)', 13),
        ('일반총원가', 13), ('일반마진(원)', 13), ('일반마진율(%)', 13),
    ]
    _xl_header(ws, cols)

    for ri, item in enumerate(data, 2):
        p, m, s = item['p'], item['m'], item['s']
        _xl_row(ws, ri, [
            p['name'], p['option_name'] or '',
            p['sale_price'], p['purchase_price_cny'], p['exchange_rate'],
            p['import_quantity'], s,
            round(m['total_cost_simple'], 1), round(m['simple_margin'], 1), round(m['simple_margin_rate'], 1),
            round(m['total_cost_general'], 1), round(m['general_margin'], 1), round(m['general_margin_rate'], 1),
        ], alt=(ri % 2 == 0))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"재고현황_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/download/history')
def download_history():
    conn = get_db()
    ins = conn.execute("""
        SELECT si.date, p.name, p.option_name, si.quantity, si.exchange_rate, si.memo
        FROM stock_in si JOIN products p ON si.product_id = p.id
        ORDER BY si.date DESC, si.id DESC
    """).fetchall()
    outs = conn.execute("""
        SELECT so.date, p.name, p.option_name, so.quantity, so.source, so.file_name, so.order_number, so.memo
        FROM stock_out so JOIN products p ON so.product_id = p.id
        ORDER BY so.date DESC, so.id DESC
    """).fetchall()
    conn.close()

    wb = openpyxl.Workbook()

    # 입고 시트
    ws_in = wb.active
    ws_in.title = '입고이력'
    ws_in.freeze_panes = 'A2'
    _xl_header(ws_in, [
        ('날짜', 12), ('상품명', 22), ('옵션명', 14),
        ('수량', 8), ('환율', 8), ('메모', 28),
    ], fill_color='059669')
    for ri, row in enumerate(ins, 2):
        _xl_row(ws_in, ri, list(row), alt=(ri % 2 == 0))

    # 출고 시트
    ws_out = wb.create_sheet('출고이력')
    ws_out.freeze_panes = 'A2'
    _xl_header(ws_out, [
        ('날짜', 12), ('상품명', 22), ('옵션명', 14),
        ('수량', 8), ('출처', 10), ('파일명', 30), ('주문번호', 18), ('메모', 20),
    ], fill_color='DC2626')
    for ri, row in enumerate(outs, 2):
        _xl_row(ws_out, ri, list(row), alt=(ri % 2 == 0))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"입출고이력_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/backup-db')
def backup_db():
    fname = f"margin_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    return send_file(DB_PATH, as_attachment=True, download_name=fname,
                     mimetype='application/octet-stream')


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)
